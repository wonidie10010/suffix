import json
import logging
import numbers
import os
import re
import sys
import warnings


PROGRESS_ACTIVE = False


def suppress_startup_noise():
    os.environ.setdefault("HF_HUB_DISABLE_PROGRESS_BARS", "1")
    os.environ.setdefault("TRANSFORMERS_NO_ADVISORY_WARNINGS", "1")
    warnings.filterwarnings(
        "ignore",
        message=r".*torch\.utils\._pytree\._register_pytree_node.*deprecated.*",
        category=FutureWarning,
    )
    warnings.filterwarnings(
        "ignore",
        message=r"Using `TRANSFORMERS_CACHE` is deprecated.*",
        category=FutureWarning,
    )
    warnings.filterwarnings("ignore", category=DeprecationWarning, module=r"matplotlib\..*")
    warnings.filterwarnings("ignore", message=r".*deprecated - use.*", category=UserWarning)
    logging.getLogger("torch.distributed.elastic.multiprocessing.redirects").setLevel(logging.ERROR)


def disable_external_progress_bars():
    os.environ.setdefault("HF_HUB_DISABLE_PROGRESS_BARS", "1")
    os.environ.setdefault("TRANSFORMERS_NO_ADVISORY_WARNINGS", "1")
    warnings.filterwarnings("ignore", message="You are using the default.*")
    try:
        from huggingface_hub.utils import disable_progress_bars
        disable_progress_bars()
    except Exception:
        pass
    try:
        from transformers.utils import logging as transformers_logging
        transformers_logging.set_verbosity_error()
        if hasattr(transformers_logging, "disable_progress_bar"):
            transformers_logging.disable_progress_bar()
    except Exception:
        pass


def console_finish_progress():
    global PROGRESS_ACTIVE
    if PROGRESS_ACTIVE:
        sys.stdout.write("\n")
        sys.stdout.flush()
        PROGRESS_ACTIVE = False


def console_safe_text(text):
    text = str(text)
    encoding = sys.stdout.encoding or "utf-8"
    return text.encode(encoding, errors="replace").decode(encoding, errors="replace")


def console_update(text):
    global PROGRESS_ACTIVE
    sys.stdout.write("\r\033[K{}".format(console_safe_text(text)))
    sys.stdout.flush()
    PROGRESS_ACTIVE = True


def format_progress_bar(percent, width=24):
    filled = int(width * percent / 100)
    return "[{}{}]".format("#" * filled, "-" * (width - filled))


def json_default(value):
    value_type = type(value)
    if value_type.__module__.startswith("torch"):
        return str(value)
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return str(value)


def dump_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=json_default)


def experiment_name_from_output_dir(output_dir):
    name = os.path.basename(os.path.normpath(output_dir or "experiment"))
    name = re.sub(r'[<>:"/\\|?*]+', "_", name).strip()
    return name or "experiment"


def experiment_method_directory_name(
        selected_advanced_method,
        selected_candidate_reranking_method=None):
    advanced_method = str(
        selected_advanced_method or "frozen_original_baseline"
    ).strip()
    candidate_method = str(
        selected_candidate_reranking_method or "none"
    ).strip()
    method_parts = []
    if advanced_method:
        method_parts.append(advanced_method)
    if candidate_method and candidate_method.lower() != "none":
        method_parts.append(candidate_method)
    if not method_parts:
        method_parts.append("frozen_original_baseline")
    name = "__".join(method_parts)
    name = re.sub(r'[<>:"/\\|?*]+', "_", name).strip(" .")
    return name or "frozen_original_baseline"


def _resolved_candidate_config_view(candidate_configs, selected_method):
    if selected_method == "CGMR_v1.2":
        return candidate_configs
    selected_key = str(selected_method).lower().replace(".", "_")
    if selected_key in candidate_configs:
        return {selected_key: candidate_configs[selected_key]}
    return {}


def _resolved_suffix_v20_config(args):
    defaults = {
        "enabled": False, "log_enabled": True,
        "layer_offsets": [0, 1, 2], "layer_weights": [1.0, 0.5, 0.25],
        "epsilon": 1e-8, "phase1_epoch": 1000, "phase1_lr": 0.01,
        "phase1_direction_weight": 0.9, "phase1_magnitude_weight": 0.1,
        "phase2_epoch": 50, "phase2_lr": 0.001,
        "phase2_direction_weight": 0.1, "phase2_magnitude_weight": 0.9,
        "score_direction_weight": 0.5, "score_magnitude_weight": 0.5,
        "prox_weight": 0.005, "range_weight": 0.001,
        "continuous_mad_multiplier": 3.0,
        "local_discrete_mad_multiplier": 3.0,
        "local_gap_jump_mad_multiplier": 3.0, "mad_epsilon": 1e-8,
        "local_min_points": 4, "normal_embedding_top_k": 10,
        "expanded_embedding_top_k": 20, "ppl_top_k": 10,
        "classifier_top_k": 10, "cumulative_min_points": 4,
        "cumulative_kappa": 0.5, "cumulative_threshold": 5.0,
        "replace_epsilon": 1e-8,
        "cumulative_max_repairs_per_trigger": 1,
        "accuracy_diagnostics_enabled": True, "classifier_enabled": False,
    }
    prefixes = {
        "enabled": "suffix_reoptimization_v2_0",
        "log_enabled": "suffix_reoptimization_v2_0_log",
    }
    resolved = {
        key: getattr(args, prefixes.get(key, "suffix_v2_0_" + key), default)
        for key, default in defaults.items()
    }
    resolved.update({
        "version": "v2.0",
        "method": "suffix_reoptimization_v2.0",
        "classifier_provider_available": False,
        "classifier_candidate_count": 0,
        "loss_metric": "direction_magnitude_joint_error_float32",
        "final_acceptance": "hard_failure_only_rollback",
    })
    return resolved


def build_resolved_config(args, timestamp, run_dir, experiment_log_path,
                          reconstruction_path, summary_excel_path,
                          total_samples, model_config_layers, model_type,
                          loaded_model_layers=None, dataset_parameters=None):
    configured_datasets = dataset_parameters or [{
        "name": os.path.splitext(os.path.basename(args.dataset_path))[0],
        "path": args.dataset_path,
        "type": args.dataset_type,
        "len": args.dataset_len,
    }]
    selected_candidate_method = getattr(
        args,
        "selected_candidate_reranking_method",
        "none",
    )
    selected_is_cgmr_v1_2 = selected_candidate_method == "CGMR_v1.2"
    resolved = {
        "run": {
            "timestamp": timestamp,
            "experiment_name": experiment_name_from_output_dir(args.output_dir),
            "seed": args.seed,
            "config": args.config,
        },
        "dataset": {
            "path": args.dataset_path,
            "type": args.dataset_type,
            "len_setting": args.dataset_len,
        },
        "datasets": configured_datasets,
        "model": {
            "base_model_name": args.base_model_name,
            "lora_model_name": args.lora_model_name,
            "model_type": model_type,
            "config_layers": model_config_layers,
            "loaded_layers": loaded_model_layers,
            "num_invert_layers": args.num_invert_layers,
            "quantization": args.quantization,
        },
        "optimization": {
            "lr": args.lr,
            "epoch": args.epoch,
            "alpha": args.alpha,
            "clip": args.clip,
            "init_method": args.init_method,
            "init_param": args.init_param,
            "optim_method": args.optim_method,
        },
        "rerank": {
            "invert_method": args.invert_method,
            "filter_nonascii": args.filter_nonascii,
            "top_k_cos": args.top_k_cos,
            "perplexity": args.perplexity,
            "top_k_ppl": args.top_k_ppl,
        },
        "advanced_method": {
            "name": args.selected_advanced_method,
            "enabled": (
                args.selected_advanced_method
                != "frozen_original_baseline"
            ),
            "suffix_version": args.suffix_version,
            "selection_rule": (
                "suffix_version selects a complete suffix method; none "
                "selects frozen_original_baseline"
            ),
        },
        "candidate_reranking_method": {
            "name": selected_candidate_method,
            "enabled": selected_candidate_method != "none",
            "cgmr_version": (
                getattr(
                    args,
                    "resolved_cgmr_version",
                    getattr(args, "cgmr_version", None),
                )
                if selected_is_cgmr_v1_2
                else getattr(args, "cgmr_version", None)
            ),
            "selection_rule": "cgmr_version selector, fallback CGMR_v1.2 > CGMR_v1.1 > CGMR_v1.0 > none",
            "execution_order": (
                "online token reranking after continuous embedding or "
                "selected upstream suffix output"
                if selected_is_cgmr_v1_2
                else "after selected_advanced_method token reranking"
            ),
        },
        "advanced_methods": {
            "suffix_reoptimization_v2_0": _resolved_suffix_v20_config(args),
            "suffix_reoptimization_v1_4_1": {
                "enabled": args.suffix_reoptimization_v1_4_1,
                "log_enabled": args.suffix_reoptimization_v1_4_1_log,
                "coarse_lr_max": args.suffix_v1_4_1_coarse_lr_max,
                "coarse_lr_min": args.suffix_v1_4_1_coarse_lr_min,
                "coarse_schedule": args.suffix_v1_4_1_coarse_schedule,
                "coarse_epoch_source": "optimization.epoch",
                "fine_epoch": args.suffix_v1_4_1_fine_epoch,
                "fine_lr_max": args.suffix_v1_4_1_fine_lr_max,
                "fine_lr_min": args.suffix_v1_4_1_fine_lr_min,
                "fine_schedule": args.suffix_v1_4_1_fine_schedule,
                "confidence_mode": args.suffix_v1_4_1_confidence_mode,
                "confidence_continuous_min": args.suffix_v1_4_1_confidence_continuous_min,
                "confidence_token_min": args.suffix_v1_4_1_confidence_token_min,
                "confidence_margin_min": args.suffix_v1_4_1_confidence_margin_min,
                "confidence_gap_max": args.suffix_v1_4_1_confidence_gap_max,
                "require_candidate_agreement": args.suffix_v1_4_1_require_candidate_agreement,
                "adaptive_z_threshold": args.suffix_v1_4_1_adaptive_z_threshold,
                "adaptive_drop_z_threshold": args.suffix_v1_4_1_adaptive_drop_z_threshold,
                "adaptive_min_std": args.suffix_v1_4_1_adaptive_min_std,
                "adaptive_min_points": args.suffix_v1_4_1_adaptive_min_points,
                "fine_window": args.suffix_v1_4_1_fine_window,
                "fine_window_decay": args.suffix_v1_4_1_fine_window_decay,
                "prox_weight": args.suffix_v1_4_1_prox_weight,
                "range_weight": args.suffix_v1_4_1_range_weight,
                "min_hidden_delta": args.suffix_v1_4_1_min_hidden_delta,
                "accuracy_tolerance": args.suffix_v1_4_1_accuracy_tolerance,
                "accept_mode": args.suffix_v1_4_1_accept_mode,
                "manifold_enabled": False,
                "manifold_weight": 0.0,
                "manifold_updates": 0,
                "loss_formula": "masked_window_hidden_loss + prox_weight * prox_loss + range_weight * range_loss",
                "optimizer_scope": "persistent SGD coarse stage + one sparse Adam fine stage",
            },
            "suffix_reoptimization_v1_4": {
                "enabled": args.suffix_reoptimization_v1_4,
                "log_enabled": args.suffix_reoptimization_v1_4_log,
                "coarse_lr_max": args.suffix_v1_4_coarse_lr_max,
                "coarse_lr_min": args.suffix_v1_4_coarse_lr_min,
                "coarse_schedule": args.suffix_v1_4_coarse_schedule,
                "coarse_epoch_source": "optimization.epoch",
                "fine_epoch": args.suffix_v1_4_fine_epoch,
                "fine_lr_max": args.suffix_v1_4_fine_lr_max,
                "fine_lr_min": args.suffix_v1_4_fine_lr_min,
                "fine_schedule": args.suffix_v1_4_fine_schedule,
                "confidence_mode": args.suffix_v1_4_confidence_mode,
                "confidence_continuous_min": args.suffix_v1_4_confidence_continuous_min,
                "confidence_token_min": args.suffix_v1_4_confidence_token_min,
                "confidence_margin_min": args.suffix_v1_4_confidence_margin_min,
                "confidence_gap_max": args.suffix_v1_4_confidence_gap_max,
                "confidence_percentile_min": args.suffix_v1_4_confidence_percentile_min,
                "confidence_min_points": args.suffix_v1_4_confidence_min_points,
                "require_candidate_agreement": args.suffix_v1_4_require_candidate_agreement,
                "adaptive_z_threshold": args.suffix_v1_4_adaptive_z_threshold,
                "adaptive_drop_z_threshold": args.suffix_v1_4_adaptive_drop_z_threshold,
                "adaptive_min_std": args.suffix_v1_4_adaptive_min_std,
                "adaptive_min_points": args.suffix_v1_4_adaptive_min_points,
                "fine_window": args.suffix_v1_4_fine_window,
                "fine_window_decay": args.suffix_v1_4_fine_window_decay,
                "prox_weight": args.suffix_v1_4_prox_weight,
                "range_weight": args.suffix_v1_4_range_weight,
                "min_hidden_delta": args.suffix_v1_4_min_hidden_delta,
                "accuracy_tolerance": args.suffix_v1_4_accuracy_tolerance,
                "accept_mode": args.suffix_v1_4_accept_mode,
                "manifold_enabled": False,
                "manifold_weight": 0.0,
                "manifold_updates": 0,
                "loss_formula": "masked_window_hidden_loss + prox_weight * prox_loss + range_weight * range_loss",
                "optimizer_scope": "persistent SGD coarse stage + one sparse Adam fine stage",
            },
            "suffix_reoptimization_v1_3": {
                "enabled": args.suffix_reoptimization_v1_3,
                "log_enabled": args.suffix_reoptimization_v1_3_log,
                "max_rounds": args.suffix_v1_3_max_rounds,
                "epoch": args.suffix_v1_3_epoch,
                "lr": args.suffix_v1_3_lr,
                "hidden_low_threshold": args.suffix_v1_3_hidden_low_threshold,
                "hidden_drop_threshold": args.suffix_v1_3_hidden_drop_threshold,
                "token_forward_low_threshold": args.suffix_v1_3_token_forward_low_threshold,
                "min_anomaly_reasons": args.suffix_v1_3_min_anomaly_reasons,
                "anomaly_detection_mode": args.suffix_v1_3_anomaly_detection_mode,
                "adaptive_z_threshold": args.suffix_v1_3_adaptive_z_threshold,
                "adaptive_drop_z_threshold": args.suffix_v1_3_adaptive_drop_z_threshold,
                "adaptive_min_std": args.suffix_v1_3_adaptive_min_std,
                "adaptive_min_points": args.suffix_v1_3_adaptive_min_points,
                "min_hidden_delta": args.suffix_v1_3_min_hidden_delta,
                "accuracy_tolerance": args.suffix_v1_3_accuracy_tolerance,
                "accept_mode": args.suffix_v1_3_accept_mode,
                "hidden_weight_mode": args.suffix_v1_3_hidden_weight_mode,
                "hidden_weight_decay": args.suffix_v1_3_hidden_weight_decay,
                "hidden_weight_floor": args.suffix_v1_3_hidden_weight_floor,
                "prox_weight": args.suffix_v1_3_prox_weight,
                "manifold_weight": args.suffix_v1_3_manifold_weight,
                "manifold_update_every": args.suffix_v1_3_manifold_update_every,
                "manifold_warmup_epoch": args.suffix_v1_3_manifold_warmup_epoch,
                "range_weight": args.suffix_v1_3_range_weight,
                "anchor_mode": args.suffix_v1_3_anchor_mode,
                "loss_formula": "weighted_hidden_loss + prox_weight * prox_loss + manifold_weight * manifold_loss + range_weight * range_loss",
                "prox_reference": "original_state_suffix",
                "optimizer_scope": "new Adam for every anchored round",
            },
            "suffix_reoptimization_v1_3_1": {
                "enabled": getattr(args, "suffix_reoptimization_v1_3_1", False),
                "max_rounds": getattr(args, "suffix_v1_3_1_max_rounds", 2),
                "epoch": getattr(args, "suffix_v1_3_1_epoch", 50),
                "lr": getattr(args, "suffix_v1_3_1_lr", 0.03),
                "hidden_low_threshold": getattr(
                    args, "suffix_v1_3_1_hidden_low_threshold", 0.50
                ),
                "hidden_drop_threshold": getattr(
                    args, "suffix_v1_3_1_hidden_drop_threshold", 0.15
                ),
                "token_forward_low_threshold": getattr(
                    args, "suffix_v1_3_1_token_forward_low_threshold", 0.50
                ),
                "min_anomaly_reasons": getattr(
                    args, "suffix_v1_3_1_min_anomaly_reasons", 2
                ),
                "anomaly_detection_mode": getattr(
                    args, "suffix_v1_3_1_anomaly_detection_mode", "adaptive"
                ),
                "adaptive_z_threshold": getattr(
                    args, "suffix_v1_3_1_adaptive_z_threshold", 1.5
                ),
                "adaptive_drop_z_threshold": getattr(
                    args, "suffix_v1_3_1_adaptive_drop_z_threshold", 1.5
                ),
                "adaptive_min_std": getattr(
                    args, "suffix_v1_3_1_adaptive_min_std", 1e-6
                ),
                "adaptive_min_points": getattr(
                    args, "suffix_v1_3_1_adaptive_min_points", 4
                ),
                "min_hidden_delta": getattr(
                    args, "suffix_v1_3_1_min_hidden_delta", 0.005
                ),
                "accuracy_tolerance": getattr(
                    args, "suffix_v1_3_1_accuracy_tolerance", 0.0
                ),
                "accept_mode": getattr(
                    args, "suffix_v1_3_1_accept_mode", "oracle_accuracy"
                ),
                "hidden_weight_mode": getattr(
                    args, "suffix_v1_3_1_hidden_weight_mode", "front_decay"
                ),
                "hidden_weight_decay": getattr(
                    args, "suffix_v1_3_1_hidden_weight_decay", 0.90
                ),
                "hidden_weight_floor": getattr(
                    args, "suffix_v1_3_1_hidden_weight_floor", 0.20
                ),
                "prox_weight": getattr(args, "suffix_v1_3_1_prox_weight", 0.005),
                "range_weight": getattr(args, "suffix_v1_3_1_range_weight", 0.001),
                "prefix_anchor_interval": "[eval_start_pos:suffix_start)",
                "optimizer_scope": "new suffix-only Adam for every anchored round",
            },
            "suffix_reoptimization_v1_2_1": {
                "enabled": args.suffix_reoptimization_v1_2_1,
                "log_enabled": args.suffix_reoptimization_v1_2_1_log,
                "max_rounds": args.suffix_v1_2_1_max_rounds,
                "epoch": args.suffix_v1_2_1_epoch,
                "lr": args.suffix_v1_2_1_lr,
                "hidden_low_threshold": args.suffix_v1_2_1_hidden_low_threshold,
                "hidden_drop_threshold": args.suffix_v1_2_1_hidden_drop_threshold,
                "token_forward_low_threshold": args.suffix_v1_2_1_token_forward_low_threshold,
                "min_anomaly_reasons": args.suffix_v1_2_1_min_anomaly_reasons,
                "anomaly_detection_mode": args.suffix_v1_2_1_anomaly_detection_mode,
                "adaptive_z_threshold": args.suffix_v1_2_1_adaptive_z_threshold,
                "adaptive_drop_z_threshold": args.suffix_v1_2_1_adaptive_drop_z_threshold,
                "adaptive_min_std": args.suffix_v1_2_1_adaptive_min_std,
                "adaptive_min_points": args.suffix_v1_2_1_adaptive_min_points,
                "min_hidden_delta": args.suffix_v1_2_1_min_hidden_delta,
                "accuracy_tolerance": args.suffix_v1_2_1_accuracy_tolerance,
                "accept_mode": args.suffix_v1_2_1_accept_mode,
                "hidden_weight_mode": args.suffix_v1_2_1_hidden_weight_mode,
                "hidden_weight_decay": args.suffix_v1_2_1_hidden_weight_decay,
                "hidden_weight_floor": args.suffix_v1_2_1_hidden_weight_floor,
                "prox_weight": args.suffix_v1_2_1_prox_weight,
                "range_weight": args.suffix_v1_2_1_range_weight,
                "manifold_enabled": False,
                "manifold_weight": 0.0,
                "manifold_updates": 0,
                "loss_formula": "weighted_hidden_loss + prox_weight * prox_loss + range_weight * range_loss",
            },
            "suffix_reoptimization_v1_2_2": {
                "enabled": getattr(args, "suffix_reoptimization_v1_2_2", False),
                "max_rounds": getattr(args, "suffix_v1_2_2_max_rounds", 2),
                "epoch": getattr(args, "suffix_v1_2_2_epoch", 50),
                "lr": getattr(args, "suffix_v1_2_2_lr", 0.03),
                "embedding_relative_mse_high_threshold": getattr(
                    args,
                    "suffix_v1_2_2_embedding_relative_mse_high_threshold",
                    1.0,
                ),
                "relative_mse_rise_threshold": getattr(
                    args, "suffix_v1_2_2_relative_mse_rise_threshold", 0.30
                ),
                "token_relative_mse_high_threshold": getattr(
                    args, "suffix_v1_2_2_token_relative_mse_high_threshold", 1.0
                ),
                "min_anomaly_reasons": getattr(
                    args, "suffix_v1_2_2_min_anomaly_reasons", 2
                ),
                "anomaly_detection_mode": getattr(
                    args, "suffix_v1_2_2_anomaly_detection_mode", "adaptive"
                ),
                "adaptive_z_threshold": getattr(
                    args, "suffix_v1_2_2_adaptive_z_threshold", 1.5
                ),
                "adaptive_rise_z_threshold": getattr(
                    args, "suffix_v1_2_2_adaptive_rise_z_threshold", 1.5
                ),
                "adaptive_min_std": getattr(
                    args, "suffix_v1_2_2_adaptive_min_std", 1e-6
                ),
                "adaptive_min_points": getattr(
                    args, "suffix_v1_2_2_adaptive_min_points", 4
                ),
                "min_relative_mse_improvement": getattr(
                    args, "suffix_v1_2_2_min_relative_mse_improvement", 0.01
                ),
                "accuracy_tolerance": getattr(
                    args, "suffix_v1_2_2_accuracy_tolerance", 0.0
                ),
                "accept_mode": getattr(
                    args, "suffix_v1_2_2_accept_mode", "oracle_accuracy"
                ),
                "hidden_weight_mode": getattr(
                    args, "suffix_v1_2_2_hidden_weight_mode", "front_decay"
                ),
                "hidden_weight_decay": getattr(
                    args, "suffix_v1_2_2_hidden_weight_decay", 0.90
                ),
                "hidden_weight_floor": getattr(
                    args, "suffix_v1_2_2_hidden_weight_floor", 0.20
                ),
                "cosine_loss_weight": getattr(
                    args, "suffix_v1_2_2_cosine_loss_weight", 0.1
                ),
                "relative_mse_loss_weight": getattr(
                    args, "suffix_v1_2_2_relative_mse_loss_weight", 0.9
                ),
                "prox_weight": getattr(args, "suffix_v1_2_2_prox_weight", 0.005),
                "range_weight": getattr(args, "suffix_v1_2_2_range_weight", 0.001),
                "gradient_trend_stats_enabled": getattr(
                    args, "suffix_v1_2_2_gradient_trend_stats_enabled", True
                ),
                "relative_mse_epsilon": 1e-8,
            },
            "suffix_v1_2_3": {
                "enabled": getattr(args, "suffix_v1_2_3", False),
                "log_enabled": getattr(
                    args, "suffix_v1_2_3_log", True
                ),
                "stage1": {
                    "metric": "relative_mse",
                    "vocab_metric": "embedding_mse",
                    "candidate_rerank_metric": "hidden_relative_mse",
                    "optimizer": "SGD recreated each epoch",
                    "lr_source": "optimization.lr",
                    "lr": args.lr,
                    "epoch_source": "optimization.epoch",
                    "epoch": args.epoch,
                    "range_weight_source": "optimization.alpha",
                    "range_weight": args.alpha,
                    "clip": args.clip,
                    "init_method": args.init_method,
                    "init_param": args.init_param,
                    "top_k_source": "rerank.top_k_cos",
                    "top_k": args.top_k_cos,
                    "embedding_search_chunk_size": 8192,
                    "filter_nonascii": args.filter_nonascii,
                    "add_perplexity": args.perplexity,
                    "top_k_ppl": args.top_k_ppl,
                    "relative_mse_epsilon": 1e-8,
                },
                "reoptimization": {
                    "source": (
                        "suffix_reoptimization_v1.2.2_unchanged"
                    ),
                    "max_rounds": getattr(
                        args, "suffix_v1_2_3_max_rounds", 2
                    ),
                    "epoch": getattr(args, "suffix_v1_2_3_epoch", 50),
                    "lr": getattr(args, "suffix_v1_2_3_lr", 0.03),
                    "embedding_relative_mse_high_threshold": getattr(
                        args,
                        "suffix_v1_2_3_embedding_relative_mse_high_threshold",
                        1.0,
                    ),
                    "relative_mse_rise_threshold": getattr(
                        args,
                        "suffix_v1_2_3_relative_mse_rise_threshold",
                        0.30,
                    ),
                    "token_relative_mse_high_threshold": getattr(
                        args,
                        "suffix_v1_2_3_token_relative_mse_high_threshold",
                        1.0,
                    ),
                    "min_anomaly_reasons": getattr(
                        args,
                        "suffix_v1_2_3_min_anomaly_reasons",
                        1,
                    ),
                    "anomaly_detection_mode": getattr(
                        args,
                        "suffix_v1_2_3_anomaly_detection_mode",
                        "adaptive",
                    ),
                    "adaptive_z_threshold": getattr(
                        args,
                        "suffix_v1_2_3_adaptive_z_threshold",
                        1.5,
                    ),
                    "adaptive_rise_z_threshold": getattr(
                        args,
                        "suffix_v1_2_3_adaptive_rise_z_threshold",
                        1.5,
                    ),
                    "adaptive_min_std": getattr(
                        args,
                        "suffix_v1_2_3_adaptive_min_std",
                        1e-6,
                    ),
                    "adaptive_min_points": getattr(
                        args,
                        "suffix_v1_2_3_adaptive_min_points",
                        4,
                    ),
                    "min_relative_mse_improvement": getattr(
                        args,
                        "suffix_v1_2_3_min_relative_mse_improvement",
                        0.01,
                    ),
                    "accuracy_tolerance": getattr(
                        args,
                        "suffix_v1_2_3_accuracy_tolerance",
                        0.0,
                    ),
                    "accept_mode": getattr(
                        args,
                        "suffix_v1_2_3_accept_mode",
                        "oracle_accuracy",
                    ),
                    "hidden_weight_mode": getattr(
                        args,
                        "suffix_v1_2_3_hidden_weight_mode",
                        "front_decay",
                    ),
                    "hidden_weight_decay": getattr(
                        args,
                        "suffix_v1_2_3_hidden_weight_decay",
                        0.90,
                    ),
                    "hidden_weight_floor": getattr(
                        args,
                        "suffix_v1_2_3_hidden_weight_floor",
                        0.20,
                    ),
                    "cosine_loss_weight": getattr(
                        args,
                        "suffix_v1_2_3_cosine_loss_weight",
                        0.1,
                    ),
                    "relative_mse_loss_weight": getattr(
                        args,
                        "suffix_v1_2_3_relative_mse_loss_weight",
                        0.9,
                    ),
                    "prox_weight": getattr(
                        args, "suffix_v1_2_3_prox_weight", 0.005
                    ),
                    "range_weight": getattr(
                        args, "suffix_v1_2_3_range_weight", 0.001
                    ),
                    "gradient_trend_stats_enabled": getattr(
                        args,
                        "suffix_v1_2_3_gradient_trend_stats_enabled",
                        True,
                    ),
                    "relative_mse_epsilon": 1e-8,
                },
            },
            "suffix_reoptimization_v1_2": {
                "enabled": args.suffix_reoptimization_v1_2,
                "log_enabled": args.suffix_reoptimization_v1_2_log,
                "max_rounds": args.suffix_v1_2_max_rounds,
                "epoch": args.suffix_v1_2_epoch,
                "lr": args.suffix_v1_2_lr,
                "hidden_low_threshold": args.suffix_v1_2_hidden_low_threshold,
                "hidden_drop_threshold": args.suffix_v1_2_hidden_drop_threshold,
                "token_forward_low_threshold": args.suffix_v1_2_token_forward_low_threshold,
                "min_anomaly_reasons": args.suffix_v1_2_min_anomaly_reasons,
                "anomaly_detection_mode": args.suffix_v1_2_anomaly_detection_mode,
                "adaptive_z_threshold": args.suffix_v1_2_adaptive_z_threshold,
                "adaptive_drop_z_threshold": args.suffix_v1_2_adaptive_drop_z_threshold,
                "adaptive_min_std": args.suffix_v1_2_adaptive_min_std,
                "adaptive_min_points": args.suffix_v1_2_adaptive_min_points,
                "min_hidden_delta": args.suffix_v1_2_min_hidden_delta,
                "accuracy_tolerance": args.suffix_v1_2_accuracy_tolerance,
                "accept_mode": args.suffix_v1_2_accept_mode,
                "hidden_weight_mode": args.suffix_v1_2_hidden_weight_mode,
                "hidden_weight_decay": args.suffix_v1_2_hidden_weight_decay,
                "hidden_weight_floor": args.suffix_v1_2_hidden_weight_floor,
                "prox_weight": args.suffix_v1_2_prox_weight,
                "manifold_weight": args.suffix_v1_2_manifold_weight,
                "manifold_update_every": args.suffix_v1_2_manifold_update_every,
                "manifold_warmup_epoch": args.suffix_v1_2_manifold_warmup_epoch,
                "range_weight": args.suffix_v1_2_range_weight,
            },
            "suffix_reoptimization_v1_1": {
                "enabled": args.suffix_reoptimization_v1_1,
                "log_enabled": args.suffix_reoptimization_v1_1_log,
                "max_rounds": args.suffix_v1_1_max_rounds,
                "epoch": args.suffix_v1_1_epoch,
                "lr": args.suffix_v1_1_lr,
                "hidden_low_threshold": args.suffix_v1_1_hidden_low_threshold,
                "hidden_drop_threshold": args.suffix_v1_1_hidden_drop_threshold,
                "token_forward_low_threshold": args.suffix_v1_1_token_forward_low_threshold,
                "min_anomaly_reasons": args.suffix_v1_1_min_anomaly_reasons,
                "min_hidden_delta": args.suffix_v1_1_min_hidden_delta,
                "accuracy_tolerance": args.suffix_v1_1_accuracy_tolerance,
                "accept_mode": args.suffix_v1_1_accept_mode,
                "hidden_weight_mode": args.suffix_v1_1_hidden_weight_mode,
                "hidden_weight_decay": args.suffix_v1_1_hidden_weight_decay,
                "hidden_weight_floor": args.suffix_v1_1_hidden_weight_floor,
                "prox_weight": args.suffix_v1_1_prox_weight,
                "manifold_weight": args.suffix_v1_1_manifold_weight,
                "manifold_update_every": args.suffix_v1_1_manifold_update_every,
                "manifold_warmup_epoch": args.suffix_v1_1_manifold_warmup_epoch,
                "range_weight": args.suffix_v1_1_range_weight,
            },
            "suffix_reoptimization_v1_0": {
                "enabled": args.suffix_reoptimization_v1_0,
                "log_enabled": args.suffix_reoptimization_v1_0_log,
                "max_rounds": args.suffix_v1_0_max_rounds,
                "epoch": args.suffix_v1_0_epoch,
                "lr": args.suffix_v1_0_lr,
                "reg_weight": args.suffix_v1_0_reg_weight,
                "hidden_low_threshold": args.suffix_v1_0_hidden_low_threshold,
                "hidden_drop_threshold": args.suffix_v1_0_hidden_drop_threshold,
                "token_forward_low_threshold": args.suffix_v1_0_token_forward_low_threshold,
                "min_anomaly_reasons": args.suffix_v1_0_min_anomaly_reasons,
                "min_hidden_delta": args.suffix_v1_0_min_hidden_delta,
                "accuracy_tolerance": args.suffix_v1_0_accuracy_tolerance,
                "accept_mode": args.suffix_v1_0_accept_mode,
            },
        },
        "candidate_reranking_methods": {
            "cgmr_v1_2": {
                "enabled": getattr(args, "cgmr_v1_2", False),
                "selected_by_selector": selected_is_cgmr_v1_2,
                "log_enabled": getattr(args, "cgmr_v1_2_log", True),
                "layer_offsets": getattr(
                    args, "cgmr_v1_2_layer_offsets", [0, 1, 2]
                ),
                "layer_weights": getattr(
                    args, "cgmr_v1_2_layer_weights", [0.5, 0.3, 0.2]
                ),
                "entropy_temperature": getattr(
                    args, "cgmr_v1_2_entropy_temperature", 0.05
                ),
                "effective_candidate_threshold": getattr(
                    args, "cgmr_v1_2_effective_candidate_threshold", 1.5
                ),
                "max_multilayer_candidates": getattr(
                    args, "cgmr_v1_2_max_multilayer_candidates", 6
                ),
                "lookahead_window": getattr(
                    args, "cgmr_v1_2_lookahead_window", 1
                ),
                "improvement_epsilon": getattr(
                    args, "cgmr_v1_2_improvement_epsilon", 1e-6
                ),
                "relative_mse_epsilon": getattr(
                    args, "cgmr_v1_2_relative_mse_epsilon", 1e-8
                ),
                "max_candidates": getattr(
                    args, "cgmr_v1_2_max_candidates", 32
                ),
                "candidate_batch_size": getattr(
                    args, "cgmr_v1_2_candidate_batch_size", 16
                ),
                "effective_layers": getattr(
                    args, "cgmr_v1_2_effective_layers", []
                ),
                "effective_weights": getattr(
                    args, "cgmr_v1_2_effective_weights", []
                ),
                "filtered_layers": getattr(
                    args, "cgmr_v1_2_filtered_layers", []
                ),
            },
            "cgmr_v1_1": {
                "enabled": getattr(args, "cgmr_v1_1", False),
                "log_enabled": getattr(args, "cgmr_v1_1_log", True),
                "layer_offsets": getattr(args, "cgmr_v1_1_layer_offsets", [0, 1, 2]),
                "layer_weights": getattr(
                    args, "cgmr_v1_1_layer_weights", [0.5, 0.3, 0.2]
                ),
                "normalization": getattr(args, "cgmr_v1_1_normalization", "zscore"),
                "consistency_weight": getattr(args, "cgmr_v1_1_consistency_weight", 0.0),
                "relative_margin_epsilon": getattr(
                    args, "cgmr_v1_1_relative_margin_epsilon", 1e-6
                ),
                "relative_margin_risk_weight": getattr(
                    args, "cgmr_v1_1_relative_margin_risk_weight", 0.7
                ),
                "low_score_risk_weight": getattr(
                    args, "cgmr_v1_1_low_score_risk_weight", 0.3
                ),
                "score_drop_risk_weight": getattr(
                    args, "cgmr_v1_1_score_drop_risk_weight", 0.0
                ),
                "low_score_threshold": getattr(args, "cgmr_v1_1_low_score_threshold", 0.80),
                "min_risk_score": getattr(args, "cgmr_v1_1_min_risk_score", 0.20),
                "risk_top_k": getattr(args, "cgmr_v1_1_risk_top_k", 20),
                "max_accepted_repairs": getattr(
                    args, "cgmr_v1_1_max_accepted_repairs", 10
                ),
                "max_candidates": getattr(args, "cgmr_v1_1_max_candidates", 32),
                "candidate_batch_size": getattr(
                    args, "cgmr_v1_1_candidate_batch_size", 16
                ),
                "min_enhanced_gain": getattr(args, "cgmr_v1_1_min_enhanced_gain", 0.05),
                "min_enhanced_margin": getattr(
                    args, "cgmr_v1_1_min_enhanced_margin", 0.05
                ),
                "max_layer_l_drop": getattr(args, "cgmr_v1_1_max_layer_l_drop", 0.02),
                "effective_layers": getattr(args, "cgmr_v1_1_effective_layers", []),
                "effective_weights": getattr(args, "cgmr_v1_1_effective_weights", []),
                "filtered_layers": getattr(args, "cgmr_v1_1_filtered_layers", []),
            },
            "cgmr_v1_0": {
                "enabled": getattr(args, "cgmr_v1_0", False),
                "log_enabled": getattr(args, "cgmr_v1_0_log", True),
                "layer_offsets": getattr(args, "cgmr_v1_0_layer_offsets", [0, 1, 2]),
                "layer_weights": getattr(
                    args, "cgmr_v1_0_layer_weights", [0.5, 0.3, 0.2]
                ),
                "normalization": getattr(args, "cgmr_v1_0_normalization", "zscore"),
                "consistency_weight": getattr(args, "cgmr_v1_0_consistency_weight", 0.0),
                "strong_margin_threshold": getattr(
                    args, "cgmr_v1_0_strong_margin_threshold", 0.01
                ),
                "weak_margin_threshold": getattr(
                    args, "cgmr_v1_0_weak_margin_threshold", 0.02
                ),
                "low_score_threshold": getattr(args, "cgmr_v1_0_low_score_threshold", 0.50),
                "weak_signals_required": getattr(
                    args, "cgmr_v1_0_weak_signals_required", 2
                ),
                "max_candidates": getattr(args, "cgmr_v1_0_max_candidates", 32),
                "candidate_batch_size": getattr(
                    args, "cgmr_v1_0_candidate_batch_size", 16
                ),
                "min_enhanced_gain": getattr(args, "cgmr_v1_0_min_enhanced_gain", 0.05),
                "min_enhanced_margin": getattr(
                    args, "cgmr_v1_0_min_enhanced_margin", 0.05
                ),
                "max_layer_l_drop": getattr(args, "cgmr_v1_0_max_layer_l_drop", 0.02),
                "max_repair_steps": getattr(args, "cgmr_v1_0_max_repair_steps", 5),
                "effective_layers": getattr(args, "cgmr_v1_0_effective_layers", []),
                "effective_weights": getattr(args, "cgmr_v1_0_effective_weights", []),
                "filtered_layers": getattr(args, "cgmr_v1_0_filtered_layers", []),
            },
        },
        "runtime": {
            "device_map": args.device_map,
            "offload_folder": args.offload_folder,
            "offload_state_dict": args.offload_state_dict,
            "max_memory": args.max_memory,
            "log_dir": args.log_dir,
            "output_dir_for_experiment_name_only": args.output_dir,
        },
        "outputs": {
            "run_dir": run_dir,
            "experiment_log": experiment_log_path,
            "reconstructions": reconstruction_path,
            **({"summary_excel": summary_excel_path} if summary_excel_path else {}),
        },
    }
    selected_advanced_method = resolved["advanced_method"]["name"]
    selected_advanced_key = str(selected_advanced_method).replace(".", "_")
    resolved["advanced_methods"] = {
        selected_advanced_key: resolved["advanced_methods"][selected_advanced_key]
    } if selected_advanced_key in resolved["advanced_methods"] else {}
    resolved["advanced_method"].pop("selection_rule", None)

    selected_candidate_method = resolved["candidate_reranking_method"]["name"]
    if selected_candidate_method == "CGMR_v1.2":
        resolved["candidate_reranking_method"][
            "selected_candidate_reranking_method"
        ] = "CGMR_v1.2"
    resolved["candidate_reranking_methods"] = _resolved_candidate_config_view(
        resolved["candidate_reranking_methods"],
        selected_candidate_method,
    )
    resolved["candidate_reranking_method"].pop("selection_rule", None)
    return resolved


def flatten_config(config, prefix=""):
    rows = []
    for key, value in config.items():
        name = "{}.{}".format(prefix, key) if prefix else key
        if isinstance(value, dict):
            rows.extend(flatten_config(value, name))
        else:
            rows.append((name, value))
    return rows


def excel_value(value):
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, default=json_default)
    if value is None:
        return ""
    return value


def mean_or_blank(values):
    values = [value for value in values if value is not None]
    if not values:
        return ""
    return float(sum(values) / len(values))


def set_column_widths(sheet, widths):
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def format_metric(value):
    if value is None:
        return "none"
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, numbers.Number):
        return "{:.6f}".format(float(value))
    if hasattr(value, "item"):
        try:
            return "{:.6f}".format(float(value.item()))
        except Exception:
            pass
    return str(value)


def style_table(sheet, max_row, max_col):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:{}{}".format(get_column_letter(max_col), max_row)


def _excel_float(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, numbers.Number):
        return float(value)
    if hasattr(value, "item"):
        try:
            return float(value.item())
        except Exception:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _format_accuracy_pair(before, after):
    before = _excel_float(before)
    after = _excel_float(after)
    if before is None or after is None:
        return ""
    return "{:.4f} → {:.4f}".format(before, after)


def _suffix_result(record):
    return (
        record.get("suffix_v1_2_3_result")
        or record.get("suffix_reoptimization_result")
        or record.get("suffix_reoptimization_v1_4_1_result")
        or record.get("suffix_reoptimization_v1_4_result")
        or record.get("suffix_reoptimization_v1_3_result")
        or record.get("suffix_reoptimization_v1_2_1_result")
        or record.get("suffix_reoptimization_v1_2_result")
        or record.get("suffix_reoptimization_v1_1_result")
        or record.get("suffix_reoptimization_v1_0_result")
        or {}
    )


def suffix_hidden_metric_view(result, stage="after"):
    """Return a version-aware view without mixing cosine and relative MSE."""
    result = result or {}
    version = result.get("version")
    metric_source = (
        result.get("reoptimization") or {}
        if version == "v1.2.3"
        else result
    )
    metrics = metric_source.get(stage) or {}
    if version in ("v1.2.2", "v1.2.3"):
        return {
            "version": version,
            "metric_system": "relative_mse",
            "lower_is_better": True,
            "mean": metrics.get("token_forward_relative_mse_mean"),
            "worst": metrics.get("token_forward_relative_mse_max"),
            "embedding_mean": metrics.get(
                "embedding_forward_relative_mse_mean"
            ),
            "embedding_worst": metrics.get(
                "embedding_forward_relative_mse_max"
            ),
        }
    return {
        "version": version,
        "metric_system": "cosine_similarity",
        "higher_is_better": True,
        "mean": metrics.get("hidden_mean"),
        "worst": metrics.get("hidden_min"),
        "embedding_mean": metrics.get("embedding_hidden_mean"),
        "embedding_worst": metrics.get("embedding_hidden_min"),
    }


def _method_metric_name(method_name):
    name = str(method_name or "").replace(".", "_")
    if name.startswith("suffix_reoptimization_"):
        return "suffix_{}".format(name[len("suffix_reoptimization_"):])
    return name


def build_stage_accuracy(record):
    summary = extract_experiment_stage_summary(record)
    selected_advanced_method = (
        record.get("selected_advanced_method")
        or "frozen_original_baseline"
    )
    selected_candidate_method = (
        record.get("selected_candidate_reranking_method")
        or (record.get("candidate_reranking_method") or {}).get("name")
        or "none"
    )
    stages = {}
    if summary["suffix_enabled"]:
        stages["pre_suffix"] = summary["baseline_accuracy"]
        stages[_method_metric_name(selected_advanced_method)] = summary["suffix_accuracy"]
    elif selected_advanced_method == "frozen_original_baseline":
        if selected_candidate_method == "CGMR_v1.2":
            stages["cgmr_v1_2_input"] = summary["baseline_accuracy"]
        else:
            stages["standalone_baseline"] = summary["baseline_accuracy"]
    else:
        stages["pre_advanced"] = summary["baseline_accuracy"]
    if summary["cgmr_enabled"]:
        stages[_method_metric_name(selected_candidate_method)] = summary["cgmr_accuracy"]
    return stages


def annotate_candidate_events_for_offline_evaluation(candidate_result,
                                                       target_token_ids):
    """Attach Ground Truth diagnostics only after CGMR has finished deciding."""
    if not isinstance(candidate_result, dict):
        return candidate_result
    targets = [int(token_id) for token_id in target_token_ids]
    summary = {
        "evaluated_event_count": 0,
        "accepted_count": 0,
        "rejected_count": 0,
        "correct_repair_count": 0,
        "incorrect_destruction_count": 0,
        "wrong_to_wrong_count": 0,
        "other_accepted_count": 0,
    }
    for event in candidate_result.get("events", []):
        position = event.get("position")
        if not isinstance(position, int) or position < 0 or position >= len(targets):
            continue
        target_token_id = targets[position]
        old_token = event.get("old_token")
        new_token = event.get("new_token")
        old_correct = old_token is not None and int(old_token) == target_token_id
        new_correct = new_token is not None and int(new_token) == target_token_id
        accepted = bool(event.get("accepted"))
        if not accepted:
            outcome = "rejected"
            summary["rejected_count"] += 1
        elif not old_correct and new_correct:
            outcome = "correct_repair"
            summary["accepted_count"] += 1
            summary["correct_repair_count"] += 1
        elif old_correct and not new_correct:
            outcome = "incorrect_destruction"
            summary["accepted_count"] += 1
            summary["incorrect_destruction_count"] += 1
        elif not old_correct and not new_correct:
            outcome = "wrong_to_wrong"
            summary["accepted_count"] += 1
            summary["wrong_to_wrong_count"] += 1
        else:
            outcome = "other_accepted"
            summary["accepted_count"] += 1
            summary["other_accepted_count"] += 1
        event["offline_evaluation"] = {
            "evaluation_only": True,
            "target_token_id": target_token_id,
            "old_token_correct": old_correct,
            "new_token_correct": new_correct,
            "outcome": outcome,
        }
        summary["evaluated_event_count"] += 1
    candidate_result["offline_evaluation_summary"] = summary
    return candidate_result


def extract_experiment_stage_summary(record):
    selected_advanced_method = (
        record.get("selected_advanced_method")
        or "frozen_original_baseline"
    )
    selected_candidate_method = (
        record.get("selected_candidate_reranking_method")
        or (record.get("candidate_reranking_method") or {}).get("name")
        or "none"
    )
    suffix_enabled = str(selected_advanced_method).startswith(
        ("suffix_reoptimization_v", "suffix_v")
    )
    cgmr_enabled = selected_candidate_method != "none"

    suffix_result = _suffix_result(record) if suffix_enabled else {}
    candidate_result = record.get("candidate_reranking_result") or {}

    if suffix_enabled:
        baseline_accuracy = _excel_float(suffix_result.get("pre_acc"))
        suffix_accuracy = _excel_float(
            suffix_result.get("post_acc", suffix_result.get("final_accuracy"))
        )
    elif cgmr_enabled:
        baseline_accuracy = _excel_float(candidate_result.get("pre_acc"))
        suffix_accuracy = None
    else:
        baseline_accuracy = _excel_float(record.get("accuracy"))
        suffix_accuracy = None

    cgmr_accuracy = None
    if cgmr_enabled:
        cgmr_accuracy = _excel_float(
            candidate_result.get("post_acc", candidate_result.get("final_accuracy"))
        )

    method_chain = str(selected_advanced_method)
    if cgmr_enabled:
        method_chain = "{} + {}".format(method_chain, selected_candidate_method)

    return {
        "selected_advanced_method": method_chain,
        "baseline_accuracy": baseline_accuracy,
        "suffix_enabled": suffix_enabled,
        "suffix_accuracy": suffix_accuracy,
        "cgmr_enabled": cgmr_enabled,
        "cgmr_accuracy": cgmr_accuracy,
    }


def _write_experiment_log_block(f, lines):
    if f.tell() > 0:
        f.write("\n")
    f.write("\n".join(lines) + "\n")
    f.flush()


def write_experiment_sample_summary(f, record, sample_number, total_samples,
                                    token_length):
    dataset = record.get("dataset") or {}
    dataset_name = dataset.get("name") or "dataset"
    dataset_sample_number = dataset.get("sample_number", sample_number)
    dataset_sample_count = dataset.get("sample_count", total_samples)
    stages = record.get("stage_accuracy") or build_stage_accuracy(record)
    summary = extract_experiment_stage_summary(record)
    lines = [
        "===== dataset {} sample {}/{} =====".format(
            dataset_name, dataset_sample_number, dataset_sample_count
        ),
        "  dataset: {}".format(dataset_name),
        "  dataset_sample: {}/{}".format(
            dataset_sample_number, dataset_sample_count
        ),
        "  global_sample: {}/{}".format(sample_number, total_samples),
        "  token_length: {}".format(token_length),
        "  selected_method: {}".format(summary["selected_advanced_method"]),
    ]
    for stage_name, accuracy in stages.items():
        lines.append("  {}_accuracy: {}".format(stage_name, format_metric(accuracy)))
    selected_candidate_method = (
        record.get("selected_candidate_reranking_method")
        or (record.get("candidate_reranking_method") or {}).get("name")
        or "none"
    )
    candidate_result = record.get("candidate_reranking_result") or {}
    if (
        selected_candidate_method == "CGMR_v1.2"
        and candidate_result.get("log_enabled", True)
    ):
        lines.extend([
            "  cgmr_name: {}".format(
                candidate_result.get("name", "CGMR_v1.2")
            ),
            "  cgmr_enabled: {}".format(
                str(bool(candidate_result.get("enabled"))).lower()
            ),
            "  cgmr_effective_layers: {}".format(
                json.dumps(
                    candidate_result.get("effective_layers", []),
                    ensure_ascii=False,
                )
            ),
            "  cgmr_effective_weights: {}".format(
                json.dumps(
                    candidate_result.get("effective_weights", []),
                    ensure_ascii=False,
                )
            ),
            "  cgmr_entropy_temperature: {}".format(
                candidate_result.get("entropy_temperature")
            ),
            "  cgmr_effective_candidate_threshold: {}".format(
                candidate_result.get("effective_candidate_threshold")
            ),
            "  cgmr_high_entropy_position_count: {}".format(
                candidate_result.get("high_entropy_position_count", 0)
            ),
            "  cgmr_multilayer_positions: {}".format(
                json.dumps(
                    candidate_result.get("multilayer_positions", []),
                    ensure_ascii=False,
                )
            ),
            "  cgmr_multilayer_accepted_positions: {}".format(
                json.dumps(
                    candidate_result.get(
                        "multilayer_accepted_positions",
                        [],
                    ),
                    ensure_ascii=False,
                )
            ),
            "  cgmr_elapsed_seconds: {}".format(
                format_metric(candidate_result.get("elapsed_seconds"))
            ),
            "  cgmr_pre_accuracy: {}".format(
                format_metric(candidate_result.get("pre_acc"))
            ),
            "  cgmr_post_accuracy: {}".format(
                format_metric(
                    candidate_result.get(
                        "post_acc",
                        candidate_result.get("final_accuracy"),
                    )
                )
            ),
        ])
    _write_experiment_log_block(f, lines)


def _write_accuracy_average_block(f, title, records):
    stage_names = []
    stage_values = {}
    for record in records:
        stages = record.get("stage_accuracy") or build_stage_accuracy(record)
        for stage_name, value in stages.items():
            if stage_name not in stage_values:
                stage_names.append(stage_name)
                stage_values[stage_name] = []
            if value is not None:
                stage_values[stage_name].append(value)
    lines = [title, "  sample_count: {}".format(len(records))]
    for stage_name in stage_names:
        lines.append(
            "  {}_average_accuracy: {}".format(
                stage_name,
                format_metric(mean_or_blank(stage_values[stage_name])),
            )
        )
    _write_experiment_log_block(f, lines)


def write_experiment_average_summary(f, records):
    grouped_records = {}
    for record in records:
        dataset_name = (record.get("dataset") or {}).get("name") or "dataset"
        grouped_records.setdefault(dataset_name, []).append(record)
    for dataset_name, dataset_records in grouped_records.items():
        _write_accuracy_average_block(
            f,
            "===== dataset {} average accuracy =====".format(dataset_name),
            dataset_records,
        )
    if len(grouped_records) > 1:
        _write_accuracy_average_block(
            f,
            "===== overall average accuracy =====",
            records,
        )


def rebuild_experiment_log(path, records):
    """Rebuild the fixed experiment log from canonical merged records."""
    records = list(records)
    with open(path, "w", encoding="utf-8") as output:
        for index, record in enumerate(records, start=1):
            token_length = record.get("token_length")
            if token_length is None:
                result = record.get("suffix_reoptimization_result") or {}
                token_length = len(result.get("final_tokens") or [])
            write_experiment_sample_summary(
                output, record, index, len(records), int(token_length)
            )
        write_experiment_average_summary(output, records)


def _suffix_events(record):
    result = _suffix_result(record)
    events = result.get("events") or []
    return events if isinstance(events, list) else []


def _triggered_suffix_events(record):
    return [event for event in _suffix_events(record) if event.get("triggered")]


def _changed_positions_count(result):
    changed_positions = result.get("changed_positions")
    if changed_positions is None:
        return 0
    if isinstance(changed_positions, (list, tuple, set)):
        return len(changed_positions)
    return 0


def _event_reason_values(events):
    reasons = []
    for event in events:
        event_reasons = event.get("anomaly_reasons")
        if isinstance(event_reasons, (list, tuple, set)):
            reasons.extend([str(reason) for reason in event_reasons if reason is not None])
        elif event_reasons is not None:
            reasons.append(str(event_reasons))
    return reasons


def _display_width(value):
    if value is None:
        return 0
    width = 0
    for char in str(value):
        width += 2 if ord(char) > 127 else 1
    return width


def _auto_fit_columns(sheet, min_widths, max_width=42):
    from openpyxl.utils import get_column_letter

    for column_index in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        values = [
            cell.value
            for cell in sheet[column_letter]
            if cell.value is not None
        ]
        best_width = max([_display_width(value) for value in values] + [0]) + 2
        best_width = max(best_width, min_widths.get(column_letter, 8))
        sheet.column_dimensions[column_letter].width = min(best_width, max_width)


def write_summary_excel(path, config, records):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "实验结果"
    sheet.sheet_view.showGridLines = True

    title_row = 1
    header_row = 3
    first_data_row = 4
    headers = [
        "样本",
        "方法",
        "首个异常点",
        "异常原因",
        "接受轮数",
        "Accuracy 前 → 后",
        "Acc gain",
        "改变位置数",
        "Anchor mode",
        "锚定位置数",
        "边界回退次数",
        "耗时 (s)",
        "manifold enabled",
        "manifold updates",
        "CGMR method",
        "CGMR layers",
        "CGMR triggers",
        "CGMR accepts",
        "CGMR Accuracy before -> after",
        "CGMR changed positions",
    ]

    sheet.cell(row=title_row, column=1, value="实验结果")
    sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(headers))
    sheet.cell(row=title_row, column=1).font = Font(bold=True, size=12)
    sheet.cell(row=title_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    triggered_sample_count = 0
    accepted_sample_count = 0
    changed_positions_total = 0
    accepted_rounds_total = 0
    elapsed_seconds_total = 0.0
    all_triggered_events = []
    before_accs = []
    after_accs = []
    cgmr_trigger_total = 0
    cgmr_accept_total = 0
    cgmr_changed_total = 0

    for row_offset, record in enumerate(records):
        result = _suffix_result(record)
        cgmr_result = record.get("candidate_reranking_result") or {}
        triggered_events = _triggered_suffix_events(record)
        first_triggered_event = triggered_events[0] if triggered_events else {}
        accepted_rounds = sum(bool(event.get("accepted")) for event in triggered_events)
        anomaly_reasons = sorted(set(_event_reason_values(triggered_events)))
        before_acc = _excel_float(result.get("pre_acc"))
        after_acc = _excel_float(result.get("post_acc", result.get("final_accuracy")))
        acc_gain = None if before_acc is None or after_acc is None else after_acc - before_acc
        changed_count = _changed_positions_count(result)
        anchor_mode = result.get("anchor_mode") or first_triggered_event.get("anchor_mode")
        anchor_count = int(result.get("anchor_count") or 0)
        boundary_rewind_count = int(result.get("boundary_rewind_count") or 0)
        elapsed_seconds = _excel_float(record.get("elapsed_seconds"))
        cgmr_trigger_count = int(
            cgmr_result.get("evaluated_position_count", cgmr_result.get("trigger_count")) or 0
        )
        cgmr_accept_count = int(cgmr_result.get("accepted_count") or 0)
        cgmr_changed_count = _changed_positions_count(cgmr_result)
        cgmr_before_acc = _excel_float(cgmr_result.get("pre_acc"))
        cgmr_after_acc = _excel_float(
            cgmr_result.get("post_acc", cgmr_result.get("final_accuracy"))
        )

        if triggered_events:
            triggered_sample_count += 1
            all_triggered_events.extend(triggered_events)
        if result.get("accepted"):
            accepted_sample_count += 1
        if before_acc is not None and after_acc is not None:
            before_accs.append(before_acc)
            after_accs.append(after_acc)
        changed_positions_total += changed_count
        accepted_rounds_total += accepted_rounds
        if elapsed_seconds is not None:
            elapsed_seconds_total += elapsed_seconds
        cgmr_trigger_total += cgmr_trigger_count
        cgmr_accept_total += cgmr_accept_count
        cgmr_changed_total += cgmr_changed_count

        row = first_data_row + row_offset
        sheet.cell(row=row, column=1, value=record.get("sample_index"))
        sheet.cell(
            row=row,
            column=2,
            value=record.get("selected_advanced_method") or result.get("name"),
        )
        sheet.cell(row=row, column=3, value=first_triggered_event.get("anomaly_position"))
        sheet.cell(row=row, column=4, value=", ".join(anomaly_reasons) if anomaly_reasons else "none")
        sheet.cell(row=row, column=5, value=accepted_rounds)
        sheet.cell(row=row, column=6, value=_format_accuracy_pair(before_acc, after_acc))
        sheet.cell(row=row, column=7, value=acc_gain)
        sheet.cell(row=row, column=8, value=changed_count)
        sheet.cell(row=row, column=9, value=anchor_mode)
        sheet.cell(row=row, column=10, value=anchor_count)
        sheet.cell(row=row, column=11, value=boundary_rewind_count)
        sheet.cell(row=row, column=12, value=elapsed_seconds)
        sheet.cell(row=row, column=13, value=result.get("manifold_enabled"))
        sheet.cell(row=row, column=14, value=result.get("manifold_updates"))
        sheet.cell(row=row, column=15, value=cgmr_result.get("name"))
        sheet.cell(
            row=row,
            column=16,
            value=", ".join(str(value) for value in cgmr_result.get("effective_layers", [])),
        )
        sheet.cell(row=row, column=17, value=cgmr_trigger_count)
        sheet.cell(row=row, column=18, value=cgmr_accept_count)
        sheet.cell(
            row=row,
            column=19,
            value=_format_accuracy_pair(cgmr_before_acc, cgmr_after_acc),
        )
        sheet.cell(row=row, column=20, value=cgmr_changed_count)

    sample_count = len(records)
    before_avg = mean_or_blank(before_accs)
    after_avg = mean_or_blank(after_accs)
    average_pair = _format_accuracy_pair(before_avg, after_avg)
    average_delta = ""
    if before_avg != "" and after_avg != "":
        average_delta = "+{:.4f}".format(float(after_avg) - float(before_avg))

    reason_values = _event_reason_values(all_triggered_events)
    unique_reasons = sorted(set(reason_values))
    if not unique_reasons:
        reason_label = "• 异常原因"
        reason_value = "none"
    elif len(unique_reasons) == 1:
        reason_label = "• 所有异常原因均为"
        reason_value = unique_reasons[0]
    else:
        reason_label = "• 异常原因包含"
        reason_value = ", ".join(unique_reasons)

    summary_start_row = first_data_row + max(sample_count, 1) + 2
    selected_method = config.get("advanced_method", {}).get("name")
    selected_candidate_method = config.get("candidate_reranking_method", {}).get("name")
    selected_config_key = str(selected_method or "").replace(".", "_")
    selected_config = config.get("advanced_methods", {}).get(
        selected_config_key, {}
    )
    manifold_summary = "enabled={}, weight={}, updates={}".format(
        selected_config.get("manifold_enabled"),
        selected_config.get("manifold_weight"),
        selected_config.get("manifold_updates"),
    )
    summary_rows = [
        ("• selected method:", selected_method, True),
        ("• anchor mode:", selected_config.get("anchor_mode"), True),
        ("• manifold:", manifold_summary, True),
        (reason_label, reason_value, True),
        ("• {}/{} 样本触发".format(triggered_sample_count, sample_count), None, False),
        ("• {}/{} 样本结果被接受".format(accepted_sample_count, sample_count), None, False),
        ("• 接受轮数总计:", str(accepted_rounds_total), True),
        ("• 平均 accuracy:", average_pair, True),
        ("• 平均提升:", average_delta, True),
        ("• changed positions 总数:", str(changed_positions_total), True),
        ("• 样本耗时总和 (s):", "{:.2f}".format(elapsed_seconds_total), True),
    ]
    summary_rows.extend([
        ("CGMR selected method:", selected_candidate_method, True),
        ("CGMR triggered positions:", str(cgmr_trigger_total), True),
        ("CGMR accepted positions:", str(cgmr_accept_total), True),
        ("CGMR changed positions:", str(cgmr_changed_total), True),
    ])
    for row_offset, (label, value, code_value) in enumerate(summary_rows):
        row = summary_start_row + row_offset
        label_cell = sheet.cell(row=row, column=1, value=label)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        if value is not None:
            value_cell = sheet.cell(row=row, column=2, value=value)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            if code_value:
                value_cell.font = Font(name="Consolas", size=10)
                value_cell.fill = PatternFill("solid", fgColor="EEF0F2")
                value_cell.border = Border(
                    left=Side(style="thin", color="DADDE1"),
                    right=Side(style="thin", color="DADDE1"),
                    top=Side(style="thin", color="DADDE1"),
                    bottom=Side(style="thin", color="DADDE1"),
                )

    pale_border = Border(bottom=Side(style="thin", color="EDEDED"))
    for row in range(header_row, first_data_row + max(sample_count, 1)):
        for column in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = pale_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in range(first_data_row, first_data_row + sample_count):
        for column in (1, 3, 5, 7, 8, 10, 11, 12, 13, 14, 16, 17, 18, 20):
            sheet.cell(row=row, column=column).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        sheet.cell(row=row, column=7).number_format = "0.000000"
        sheet.cell(row=row, column=12).number_format = "0.00"

    _auto_fit_columns(sheet, {
        "A": 10,
        "B": 34,
        "C": 14,
        "D": 35,
        "E": 12,
        "F": 24,
        "G": 14,
        "H": 16,
        "I": 14,
        "J": 14,
        "K": 16,
        "L": 14,
        "M": 18,
        "N": 18,
        "O": 42,
        "P": 18,
        "Q": 14,
        "R": 14,
        "S": 28,
        "T": 20,
    })
    sheet.row_dimensions[title_row].height = 24

    workbook.save(path)


def log_line(f, text, console=True):
    if console:
        console_finish_progress()
        print(console_safe_text(text), flush=True)
    f.write(text + "\n")
    f.flush()


def log_section(f, title):
    log_line(f, "", console=False)
    log_line(f, "===== {} =====".format(title), console=False)


def log_kv(f, key, value):
    log_line(f, "  {}: {}".format(key, value), console=False)
