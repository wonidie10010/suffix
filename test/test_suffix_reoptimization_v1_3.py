import inspect
import os
import tempfile
import types
import unittest
from unittest import mock

import torch

import suffix_optimization_methods.method_versions.suffix_reoptimization_v1_3 as v13
from experiment_outputs import _suffix_result, build_resolved_config, write_summary_excel
from invert import load_config, normalize_suffix_version, select_advanced_method
from suffix_optimization_methods.method_versions.suffix_reoptimization_v1_2 import SuffixReoptimizationV12Config
from suffix_optimization_methods.method_versions.suffix_reoptimization_v1_2_1 import SuffixReoptimizationV121Config
from suffix_optimization_methods.method_versions.suffix_reoptimization_v1_3 import (
    METHOD_NAME,
    SuffixReoptimizationV13Config,
    _build_anchored_candidate,
    _merge_suffix,
    _optimize_suffix,
    _resolve_effective_suffix_start,
    _select_next_state,
    run_suffix_reoptimization_v1_3,
)


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


class SuffixReoptimizationV13Tests(unittest.TestCase):
    def embedding_layer(self):
        layer = torch.nn.Embedding(12, 3)
        with torch.no_grad():
            layer.weight.copy_(torch.arange(36, dtype=torch.float32).view(12, 3))
        return layer

    def test_v13_defaults_preserve_all_v12_fields(self):
        old_config = SuffixReoptimizationV12Config()
        new_config = SuffixReoptimizationV13Config()
        for field in old_config.__dataclass_fields__:
            self.assertEqual(getattr(old_config, field), getattr(new_config, field), field)
        self.assertEqual("anchor_stable_prefix", new_config.anchor_mode)

    def test_anchor_off_delegates_to_v12_without_behavior_changes(self):
        expected_embedding = torch.tensor([[[1.0], [2.0]]])
        v12_result = {
            "name": "suffix_reoptimization_v1.2",
            "enabled": True,
            "skipped": False,
            "before": {
                "embedding_hidden_mean": 0.7,
                "embedding_hidden_min": 0.6,
            },
            "after": {
                "embedding_hidden_mean": 0.8,
                "embedding_hidden_min": 0.7,
            },
            "events": [{
                "triggered": True,
                "anomaly_position": 1,
                "before_hidden_mean": 0.6,
                "before_hidden_min": 0.5,
                "candidate_hidden_mean": 0.7,
                "candidate_hidden_min": 0.6,
                "before_anomaly_count": 1,
                "candidate_anomaly_count": 0,
                "optimization": {"manifold_updates": 1},
            }],
        }
        config = SuffixReoptimizationV13Config(
            enabled=True,
            anchor_mode="anchor_off",
            log_enabled=True,
        )
        with mock.patch.object(
            v13.v12,
            "run_suffix_reoptimization_v1_2",
            return_value=(expected_embedding, v12_result),
        ) as delegated:
            actual_embedding, result = run_suffix_reoptimization_v1_3(
                None, None, expected_embedding, None, None, 0, None, None, None,
                config,
            )

        self.assertTrue(torch.equal(expected_embedding, actual_embedding))
        self.assertEqual(METHOD_NAME, result["name"])
        self.assertEqual("anchor_off", result["anchor_mode"])
        self.assertEqual([], result["events"][0]["anchored_positions"])
        delegated.assert_called_once()
        delegated_config = delegated.call_args.kwargs["config"]
        self.assertFalse(delegated_config.log_enabled)
        for field in SuffixReoptimizationV12Config.__dataclass_fields__:
            if field == "log_enabled":
                continue
            self.assertEqual(getattr(config, field), getattr(delegated_config, field))

    def test_full_prefix_uses_current_tokens_and_only_valid_prefix_positions(self):
        layer = self.embedding_layer()
        current = torch.full((1, 5, 3), -10.0)
        current_tokens = [7, 6, 5, 4, 3]
        anchored, anchored_positions, fixed_positions = _build_anchored_candidate(
            current,
            current_tokens,
            layer,
            effective_suffix_start=4,
            attention_mask=torch.tensor([[1, 1, 0, 1, 1]]),
            eval_start_pos=1,
            anchor_mode="anchor_full_prefix",
        )

        self.assertEqual([1, 3], anchored_positions)
        self.assertEqual([0], fixed_positions)
        for pos in (0, 1, 3):
            self.assertTrue(torch.equal(anchored[0, pos], layer.weight[current_tokens[pos]]))
        self.assertTrue(torch.equal(anchored[0, 2], current[0, 2]))
        self.assertTrue(torch.equal(anchored[0, 4], current[0, 4]))
        self.assertNotIn("total_input_ids", inspect.signature(_build_anchored_candidate).parameters)
        self.assertNotIn("total_input_ids", inspect.getsource(_build_anchored_candidate))

    def test_stable_prefix_rewinds_to_earliest_existing_instability(self):
        effective, unstable, rewound = _resolve_effective_suffix_start(
            detected_anomaly_start=6,
            anchor_mode="anchor_stable_prefix",
            eval_start_pos=1,
            full_anomalies=[{"position": 3}, {"position": 6}],
            rejected_boundaries={4},
            accepted_changed_positions={5},
            valid_positions=list(range(8)),
        )
        self.assertEqual(3, effective)
        self.assertEqual([3, 4, 5], unstable)
        self.assertTrue(rewound)

    def test_suffix_is_unchanged_before_optimization_and_invalid_positions_are_not_anchored(self):
        layer = self.embedding_layer()
        current = torch.randn(1, 6, 3)
        anchored, anchored_positions, _ = _build_anchored_candidate(
            current,
            [1, 2, 3, 4, 5, 6],
            layer,
            effective_suffix_start=4,
            attention_mask=torch.tensor([[1, 1, 0, 1, 1, 0]]),
            eval_start_pos=1,
            anchor_mode="anchor_stable_prefix",
        )
        self.assertEqual([1, 3], anchored_positions)
        self.assertTrue(torch.equal(current[:, 4:, :], anchored[:, 4:, :]))
        self.assertTrue(torch.equal(current[:, 2, :], anchored[:, 2, :]))

    def test_anchored_prefix_is_detached_and_suffix_receives_gradient(self):
        layer = self.embedding_layer()
        current = torch.randn(1, 5, 3, requires_grad=True)
        anchored, _, _ = _build_anchored_candidate(
            current,
            [1, 2, 3, 4, 5],
            layer,
            effective_suffix_start=3,
            attention_mask=torch.ones(1, 5),
            eval_start_pos=1,
            anchor_mode="anchor_full_prefix",
        )
        suffix_param = anchored[:, 3:, :].clone().detach().requires_grad_(True)
        merged = _merge_suffix(anchored, 3, suffix_param)
        merged.sum().backward()

        self.assertFalse(anchored.requires_grad)
        self.assertIsNone(current.grad)
        self.assertIsNotNone(suffix_param.grad)
        self.assertTrue(torch.all(suffix_param.grad != 0))

    def test_prox_reference_is_original_state_suffix(self):
        layer = self.embedding_layer()
        original = torch.tensor(
            [[[1.0, 0.5, 0.25], [0.9, 0.4, 0.2], [0.8, 0.3, 0.1]]]
        )
        anchored = original.clone()
        anchored[:, 0, :] = layer.weight[2]
        prox_reference = original[:, 1:, :].clone()
        config = SuffixReoptimizationV13Config(
            epoch=1,
            lr=0.01,
            manifold_weight=0.0,
            range_weight=0.0,
        )
        with mock.patch.object(
            v13,
            "_forward_embedding_hidden",
            side_effect=lambda model, current, attention, layer_id, hooks: current,
        ):
            _, summary = _optimize_suffix(
                None,
                anchored,
                anchored.clone(),
                None,
                0,
                None,
                1,
                config,
                layer,
                prox_reference_suffix=prox_reference,
            )

        self.assertEqual("original_state_suffix", summary["prox_reference_source"])
        self.assertEqual(0.0, summary["prox_loss_start"])
        self.assertTrue(summary["optimizer_recreated_for_round"])
        source = inspect.getsource(_optimize_suffix)
        self.assertIn("float(config.manifold_weight) * manifold_loss", source)
        self.assertIn("float(config.range_weight) * range_loss", source)

    def test_rejected_candidate_cannot_pollute_current_state(self):
        original_embedding = torch.tensor([[[1.0], [2.0], [3.0]]])
        candidate_embedding = torch.tensor([[[9.0], [8.0], [7.0]]])
        original_metrics = {"accuracy": 0.5}
        candidate_metrics = {"accuracy": 0.4}
        embedding, tokens, metrics = _select_next_state(
            False,
            original_embedding,
            [1, 2, 3],
            original_metrics,
            candidate_embedding,
            [9, 8, 7],
            candidate_metrics,
        )
        candidate_embedding.add_(100)

        self.assertTrue(torch.equal(embedding, original_embedding))
        self.assertEqual([1, 2, 3], tokens)
        self.assertIs(metrics, original_metrics)

    def test_accepted_candidate_keeps_anchored_prefix_and_optimized_suffix(self):
        original_embedding = torch.zeros(1, 4, 1)
        candidate_embedding = torch.tensor([[[5.0], [6.0], [7.0], [8.0]]])
        embedding, tokens, metrics = _select_next_state(
            True,
            original_embedding,
            [1, 1, 1, 1],
            {"accuracy": 0.1},
            candidate_embedding,
            [5, 6, 7, 8],
            {"accuracy": 0.2},
        )

        self.assertTrue(torch.equal(embedding, candidate_embedding))
        self.assertEqual([5, 6, 7, 8], tokens)
        self.assertEqual(0.2, metrics["accuracy"])

    def test_each_round_reanchors_from_latest_accepted_tokens(self):
        layer = self.embedding_layer()
        current = torch.zeros(1, 4, 3)
        first_tokens = [1, 2, 3, 4]
        first, _, _ = _build_anchored_candidate(
            current, first_tokens, layer, 3, torch.ones(1, 4), 1,
            "anchor_stable_prefix",
        )
        accepted_tokens = [1, 7, 8, 4]
        second, _, _ = _build_anchored_candidate(
            first, accepted_tokens, layer, 3, torch.ones(1, 4), 1,
            "anchor_stable_prefix",
        )

        self.assertTrue(torch.equal(second[0, 1], layer.weight[7]))
        self.assertTrue(torch.equal(second[0, 2], layer.weight[8]))
        self.assertFalse(torch.equal(second[0, 1], layer.weight[2]))

    def test_selector_resolved_config_unified_result_and_excel_report_v13(self):
        disabled = types.SimpleNamespace(enabled=False)
        v13_config = SuffixReoptimizationV13Config(enabled=True)
        v121_config = SuffixReoptimizationV121Config(enabled=True)
        v12_config = SuffixReoptimizationV12Config(enabled=True)
        for alias in (
            "1.3",
            "v1.3",
            "suffix_reoptimization_v1.3",
            "suffix_reoptimization_v1_3",
        ):
            self.assertEqual("v1.3", normalize_suffix_version(alias))
        self.assertEqual(
            METHOD_NAME,
            select_advanced_method(
                None,
                v121_config,
                v12_config,
                disabled,
                disabled,
                suffix_reopt_v1_3_config=v13_config,
            ),
        )
        self.assertEqual(
            "suffix_reoptimization_v1.2.1",
            select_advanced_method(
                "v1.2.1",
                v121_config,
                v12_config,
                disabled,
                disabled,
                suffix_reopt_v1_3_config=v13_config,
            ),
        )

        config_path = os.path.join(
            ROOT, "suffix_optimization_methods", "configs", "advanced_methods.json"
        )
        merged = load_config(config_path)
        merged.update({
            "config": config_path,
            "suffix_version": "v1.3",
            "output_dir": "vscode_deml_inversion",
            "seed": 0,
            "dataset_path": "skytrax-reviews-dataset/data/airport.csv",
            "dataset_type": "local",
            "dataset_len": 5,
            "base_model_name": "Qwen/Qwen2.5-1.5B",
            "lora_model_name": None,
            "num_invert_layers": 27,
            "quantization": "none",
            "lr": 0.1,
            "epoch": 1000,
            "alpha": 0.001,
            "clip": True,
            "init_method": "uniform",
            "init_param": 0.1,
            "optim_method": "cosine",
            "invert_method": "cosine",
            "filter_nonascii": True,
            "top_k_cos": 10,
            "perplexity": True,
            "top_k_ppl": 10,
            "selected_advanced_method": METHOD_NAME,
            "device_map": "manual",
            "offload_folder": None,
            "offload_state_dict": False,
            "max_memory": None,
            "log_dir": "results/invert_timestamp_runs",
        })
        args = types.SimpleNamespace(**merged)
        resolved = build_resolved_config(
            args,
            "timestamp",
            "run_dir",
            "experiment.log",
            "reconstructions.jsonl",
            "summary.xlsx",
            5,
            28,
            "qwen2",
            28,
        )
        self.assertEqual(METHOD_NAME, resolved["advanced_method"]["name"])
        self.assertEqual("v1.3", resolved["advanced_method"]["suffix_version"])
        self.assertNotIn("actual_len", resolved["dataset"])
        self.assertEqual(5, resolved["datasets"][0]["len"])
        self.assertEqual(
            "anchor_stable_prefix",
            resolved["advanced_methods"]["suffix_reoptimization_v1_3"]["anchor_mode"],
        )
        self.assertIn(
            "manifold_weight * manifold_loss",
            resolved["advanced_methods"]["suffix_reoptimization_v1_3"]["loss_formula"],
        )

        result = {
            "name": METHOD_NAME,
            "anchor_mode": "anchor_stable_prefix",
            "anchor_count": 3,
            "boundary_rewind_count": 1,
            "pre_acc": 0.25,
            "post_acc": 0.50,
            "accepted": True,
            "changed_positions": [2, 3],
            "manifold_enabled": True,
            "manifold_updates": 2,
            "events": [{
                "triggered": True,
                "accepted": True,
                "anomaly_position": 2,
                "anomaly_reasons": ["token_forward_adaptive_low"],
            }],
        }
        record = {
            "sample_index": 0,
            "elapsed_seconds": 12.5,
            "selected_advanced_method": METHOD_NAME,
            "suffix_reoptimization_v1_3_result": result,
            "suffix_reoptimization_result": result,
        }
        self.assertIs(result, _suffix_result(record))

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, "summary.xlsx")
            write_summary_excel(path, resolved, [record])
            workbook = load_workbook(path, data_only=True, read_only=True)
            values = [
                cell
                for row in workbook.active.iter_rows(values_only=True)
                for cell in row
                if cell is not None
            ]
            workbook.close()
        self.assertIn(METHOD_NAME, values)
        self.assertIn("anchor_stable_prefix", values)
        self.assertIn("边界回退次数", values)


if __name__ == "__main__":
    unittest.main()
