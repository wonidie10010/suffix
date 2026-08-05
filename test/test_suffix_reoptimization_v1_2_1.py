import inspect
import os
import tempfile
import types
import unittest
from unittest import mock

import torch

import suffix_optimization_methods.method_versions.suffix_reoptimization_v1_2_1 as v121
from experiment_outputs import build_resolved_config, write_summary_excel
from invert import load_config, normalize_suffix_version, select_advanced_method
from suffix_optimization_methods.method_versions.suffix_reoptimization_v1_2 import (
    SuffixReoptimizationV12Config,
    _find_anomalies as find_v12_anomalies,
)
from suffix_optimization_methods.method_versions.suffix_reoptimization_v1_2_1 import (
    METHOD_NAME,
    SuffixReoptimizationV121Config,
    _find_anomalies,
    _optimize_suffix,
    run_suffix_reoptimization_v1_2_1,
)


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


class SuffixReoptimizationV121Tests(unittest.TestCase):
    def adaptive_config(self, **overrides):
        values = {
            "anomaly_detection_mode": "adaptive",
            "adaptive_z_threshold": 1.5,
            "adaptive_drop_z_threshold": 1.5,
            "adaptive_min_std": 1e-6,
            "adaptive_min_points": 4,
        }
        values.update(overrides)
        return SuffixReoptimizationV121Config(**values)

    def test_v12_adaptive_detection_behavior_is_preserved(self):
        cases = [
            (
                [0.95, 0.94, 0.95, 0.94, 0.95, 0.94],
                [0.92, 0.91, 0.90, 0.35, 0.91, 0.90],
                0,
            ),
            (
                [0.95, 0.94, 0.95, 0.94, 0.95, 0.94, 0.95],
                [0.95, 0.94, 0.93, 0.70, 0.69, 0.68, 0.67],
                0,
            ),
        ]
        old_config = SuffixReoptimizationV12Config()
        new_config = SuffixReoptimizationV121Config()

        for embedding_scores, token_scores, scan_pos in cases:
            self.assertEqual(
                find_v12_anomalies(embedding_scores, token_scores, scan_pos, old_config),
                _find_anomalies(embedding_scores, token_scores, scan_pos, new_config),
            )

    def test_smooth_sequence_does_not_trigger(self):
        scores = [0.91, 0.90, 0.92, 0.91, 0.90, 0.92]
        self.assertEqual([], _find_anomalies(scores, scores, 0, self.adaptive_config()))

    def test_token_forward_adaptive_low_triggers_on_local_valley(self):
        anomalies = _find_anomalies(
            [0.95, 0.94, 0.95, 0.94, 0.95, 0.94],
            [0.92, 0.91, 0.90, 0.35, 0.91, 0.90],
            0,
            self.adaptive_config(),
        )

        self.assertEqual(3, anomalies[0]["position"])
        self.assertIn("token_forward_adaptive_low", anomalies[0]["reasons"])

    def test_token_forward_drop_triggers_on_sudden_drop(self):
        anomalies = _find_anomalies(
            [0.95, 0.94, 0.95, 0.94, 0.95, 0.94, 0.95],
            [0.95, 0.94, 0.93, 0.70, 0.69, 0.68, 0.67],
            0,
            self.adaptive_config(),
        )

        self.assertEqual(3, anomalies[0]["position"])
        self.assertIn("token_forward_drop", anomalies[0]["reasons"])

    def test_scan_pos_skips_rejected_position(self):
        anomalies = _find_anomalies(
            [0.95, 0.94, 0.95, 0.94, 0.95, 0.94],
            [0.92, 0.91, 0.35, 0.91, 0.90, 0.91],
            3,
            self.adaptive_config(),
        )
        self.assertEqual([], anomalies)

    def test_threshold_mode_remains_compatible(self):
        config = SuffixReoptimizationV121Config(
            anomaly_detection_mode="threshold",
            hidden_low_threshold=0.50,
            hidden_drop_threshold=0.30,
            token_forward_low_threshold=0.75,
            min_anomaly_reasons=1,
        )
        anomalies = _find_anomalies(
            [0.95, 0.94, 0.93, 0.92],
            [0.90, 0.88, 0.70, 0.89],
            0,
            config,
        )

        self.assertEqual(1, len(anomalies))
        self.assertEqual(["token_forward_low"], anomalies[0]["reasons"])
        self.assertEqual("threshold", anomalies[0]["anomaly_detection_mode"])

    def test_non_manifold_defaults_match_v12(self):
        old_config = SuffixReoptimizationV12Config()
        new_config = SuffixReoptimizationV121Config()
        fields = (
            "max_rounds",
            "epoch",
            "lr",
            "hidden_low_threshold",
            "hidden_drop_threshold",
            "token_forward_low_threshold",
            "min_anomaly_reasons",
            "min_hidden_delta",
            "accuracy_tolerance",
            "accept_mode",
            "anomaly_detection_mode",
            "adaptive_z_threshold",
            "adaptive_drop_z_threshold",
            "adaptive_min_std",
            "adaptive_min_points",
            "hidden_weight_mode",
            "hidden_weight_decay",
            "hidden_weight_floor",
            "prox_weight",
            "range_weight",
        )
        for field in fields:
            self.assertEqual(getattr(old_config, field), getattr(new_config, field), field)
        self.assertFalse(hasattr(new_config, "manifold_weight"))

    def test_loss_omits_manifold_term(self):
        source = inspect.getsource(_optimize_suffix)
        self.assertNotIn("manifold_loss", source)
        self.assertNotIn("config.manifold_weight", source)
        self.assertNotIn("_nearest_vocab_embeddings(", source)
        self.assertIn("float(config.prox_weight) * prox_loss", source)
        self.assertIn("float(config.range_weight) * range_loss", source)

    def test_optimization_never_calls_nearest_vocab_lookup(self):
        config = SuffixReoptimizationV121Config(epoch=2, range_weight=0.0)
        input_embed = torch.tensor(
            [[[1.0, 0.0], [0.8, 0.2], [0.2, 0.8]]],
            dtype=torch.float32,
        )
        target_hidden = torch.tensor(
            [[[1.0, 0.0], [1.0, 0.0], [1.0, 0.0]]],
            dtype=torch.float32,
        )
        embed_layer = types.SimpleNamespace(weight=torch.eye(2, dtype=torch.float32))

        with mock.patch.object(
            v121,
            "_nearest_vocab_embeddings",
            side_effect=AssertionError("nearest-vocabulary lookup must not run"),
            create=True,
        ) as nearest_lookup, mock.patch.object(
            v121,
            "_forward_embedding_hidden",
            side_effect=lambda model, current, attention, layer_id, hooks: current,
        ):
            _, summary = _optimize_suffix(
                model=None,
                input_embed=input_embed,
                target_hidden_state=target_hidden,
                attention_mask=None,
                layer_id=0,
                register_layer_hooks=None,
                suffix_start=1,
                config=config,
                embed_layer=embed_layer,
            )

        nearest_lookup.assert_not_called()
        self.assertFalse(summary["manifold_enabled"])
        self.assertEqual(0.0, summary["manifold_weight"])
        self.assertEqual(0, summary["manifold_updates"])

    def test_selector_supports_v121_aliases_and_v12_rollback(self):
        v121_config = SuffixReoptimizationV121Config(enabled=True)
        v12_config = SuffixReoptimizationV12Config(enabled=True)
        disabled = types.SimpleNamespace(enabled=False)

        for alias in (
            "1.2.1",
            "v1.2.1",
            "suffix_reoptimization_v1.2.1",
            "suffix_reoptimization_v1_2_1",
        ):
            self.assertEqual("v1.2.1", normalize_suffix_version(alias))
        self.assertEqual(
            "suffix_reoptimization_v1.2.1",
            select_advanced_method(None, v121_config, v12_config, disabled, disabled, disabled),
        )
        self.assertEqual(
            "suffix_reoptimization_v1.2",
            select_advanced_method("v1.2", v121_config, v12_config, disabled, disabled, disabled),
        )

    def test_resolved_config_and_result_report_v121(self):
        config_path = os.path.join(
            ROOT, "suffix_optimization_methods", "configs", "advanced_methods.json"
        )
        merged = load_config(config_path)
        merged.update({
            "config": config_path,
            "output_dir": "vscode_deml_inversion",
            "seed": 0,
            "dataset_path": "skytrax-reviews-dataset/data/airport.csv",
            "dataset_type": "local",
            "dataset_len": 5,
            "base_model_name": "Qwen/Qwen2.5-1.5B",
            "lora_model_name": None,
            "num_invert_layers": 27,
            "quantization": "none",
            "lr": 0.1,
            "epoch": 1000,
            "alpha": 0.001,
            "clip": True,
            "init_method": "uniform",
            "init_param": 0.1,
            "optim_method": "cosine",
            "invert_method": "cosine",
            "filter_nonascii": True,
            "top_k_cos": 10,
            "perplexity": True,
            "top_k_ppl": 10,
            "selected_advanced_method": METHOD_NAME,
            "suffix_version": "v1.2.1",
            "device_map": "manual",
            "offload_folder": None,
            "offload_state_dict": False,
            "max_memory": None,
            "log_dir": "results/invert_timestamp_runs",
        })
        args = types.SimpleNamespace(**merged)
        resolved = build_resolved_config(
            args,
            "timestamp",
            "run_dir",
            "experiment.log",
            "reconstructions.jsonl",
            "summary.xlsx",
            5,
            28,
            "qwen2",
            28,
        )
        method_config = resolved["advanced_methods"]["suffix_reoptimization_v1_2_1"]

        self.assertEqual(METHOD_NAME, resolved["advanced_method"]["name"])
        self.assertEqual("v1.2.1", resolved["advanced_method"]["suffix_version"])
        self.assertFalse(method_config["manifold_enabled"])
        self.assertEqual(0.0, method_config["manifold_weight"])
        self.assertEqual(0, method_config["manifold_updates"])

        _, result = run_suffix_reoptimization_v1_2_1(
            None, None, None, None, None, None, None, None, None,
            SuffixReoptimizationV121Config(enabled=False),
        )
        self.assertEqual(METHOD_NAME, result["name"])
        self.assertFalse(result["manifold_enabled"])
        self.assertEqual(0, result["manifold_updates"])

    def test_excel_summary_identifies_v121_and_manifold_ablation(self):
        from openpyxl import load_workbook

        config = {
            "advanced_method": {"name": METHOD_NAME},
            "advanced_methods": {
                "suffix_reoptimization_v1_2_1": {
                    "manifold_enabled": False,
                    "manifold_weight": 0.0,
                    "manifold_updates": 0,
                }
            },
        }
        records = [{
            "sample_index": 0,
            "elapsed_seconds": 12.5,
            "selected_advanced_method": METHOD_NAME,
            "suffix_reoptimization_result": {
                "name": METHOD_NAME,
                "pre_acc": 0.25,
                "post_acc": 0.50,
                "accepted": True,
                "manifold_enabled": False,
                "manifold_updates": 0,
                "changed_positions": [2, 3],
                "events": [{
                    "triggered": True,
                    "accepted": True,
                    "anomaly_position": 2,
                    "anomaly_reasons": ["token_forward_adaptive_low"],
                }],
            },
        }]
        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, "summary.xlsx")
            write_summary_excel(path, config, records)
            workbook = load_workbook(path, data_only=True, read_only=True)
            values = [
                cell
                for row in workbook.active.iter_rows(values_only=True)
                for cell in row
                if cell is not None
            ]
            workbook.close()

        self.assertIn(METHOD_NAME, values)
        self.assertIn("token_forward_adaptive_low", values)
        self.assertIn("manifold enabled", values)
        self.assertIn(False, values)
        self.assertIn(12.5, values)


if __name__ == "__main__":
    unittest.main()
